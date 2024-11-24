import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter


class MovieDataExcel:
    def __init__(self, file_name='热辣滚烫电影数据.xlsx'):
        self.file_name = file_name
        self.wb = None
        self.ws = None
        self.border = None
        self.current_index = 1
        self.create_template()

    def create_template(self):
        # 判断文件是否存在
        if not os.path.exists(self.file_name):
            # 如果不存在，则创建一个新的 Workbook
            # 创建工作簿和工作表
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "电影《热辣滚烫》"

            # 设置标题样式
            title_font = Font(size=14, bold=True)
            fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            self.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

            # 合并单元格并设置标题样式
            self.ws.merge_cells("A1:J1")  # 标题需要涵盖所有列，从A到J
            self.ws.row_dimensions[1].height = 50  # 设置第一行的高度为30
            self.ws["A1"].value = "电影《热辣滚烫》"  # 设置标题内容
            self.ws["A1"].font = title_font
            self.ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            self.ws["A1"].fill = fill_yellow
            self.ws["A1"].border = self.border

            # 设置表头
            headers = ["序号", "日期", "发布文案", "链接", "视频时长", "点赞", "评论", "收藏", "转发", "置顶/高赞评论1-3个"]
            self.ws.append(headers)

            # 设置表头样式
            for cell in self.ws[2]:  # 表头在第2行
                cell.font = title_font
                cell.border = self.border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 设置列宽
            column_widths = [5, 15, 30, 30, 10, 10, 10, 10, 10, 50]
            for i, width in enumerate(column_widths, start=1):
                self.ws.column_dimensions[get_column_letter(i)].width = width

            # 设置单元格样式
            for row in self.ws.iter_rows(min_row=3, max_row=self.ws.max_row, min_col=1, max_col=10):  # 数据从第3行开始
                for cell in row:
                    cell.border = self.border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # 保存为 Excel 文件
            self.wb.save(self.file_name)
        else:
            # 如果存在，则加载 Excel 文件
            self.wb = load_workbook(self.file_name)
            self.ws = self.wb.active
            # 从已有数据中获取当前的序号
            self.current_index = self.ws.max_row  # 以当前行数作为序号

    def insert_data(self, desc, create_date, video_duration, video_url, digg_count, comment_count, collect_count, share_count, top_comments):
        new_data = [self.current_index, create_date, desc, video_url, video_duration, digg_count, comment_count, collect_count, share_count]
        if top_comments:
            high_comments = ';\n'.join([f"来自: {comment['user']['nickname']}, 评论: {comment['text']}, 点赞数: {comment['digg_count']}" for comment in top_comments])
            new_data.append(high_comments)
        else:
            new_data.append("无高赞评论")

        self.ws.append(new_data)
        self.current_index += 1  # 序号递增

        # 设置单元格样式
        for row in self.ws.iter_rows(min_row=3, max_row=self.ws.max_row, min_col=1, max_col=10):  # 数据从第3行开始
            for cell in row:
                cell.border = self.border
                cell.alignment = Alignment(horizontal="center", vertical="center")
        self.wb.save(self.file_name)


if __name__ == "__main__":
    movie_excel = MovieDataExcel()
    movie_excel.insert_data(
        ['1', '2024-11-01', '发布文案示例', 'http://example.com', '00:20', '100', '20', '10', '5', '高赞评论示例'])
